from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_GLOB = "*保研去向*.xlsx"
GENERATED_SOURCE = "past-admission-2021-cohort"


def stable_id(prefix: str, *parts: str) -> str:
    raw = "|".join(parts).encode("utf-8")
    digest = hashlib.sha1(raw).hexdigest()[:12]
    return f"{prefix}_{digest}"


def file_digest(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def format_code(value: object) -> str:
    text = clean(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(\.0)?", text):
        text = text.split(".", 1)[0]
    if re.fullmatch(r"\d{5}", text):
        return text.zfill(6)
    return text


def strip_code_prefix(value: str) -> str:
    return re.sub(r"^\d+\|", "", value).strip()


def cohort_from_file(path: Path) -> tuple[str, int | None, int | None]:
    match = re.search(r"(\d{2})级", path.name)
    if not match:
        return "", None, None
    class_year = 2000 + int(match.group(1))
    return f"{match.group(1)}级", class_year, class_year + 4


def field_for_header(header: str) -> str:
    normalized = re.sub(r"\s+", "", header)
    if normalized in {"学院（系）名称", "学院(系)名称"}:
        return "undergraduateCollege"
    if normalized == "专业名称":
        return "undergraduateMajor"
    if normalized == "推荐类型":
        return "recommendationType"
    if "拟录取院校" in normalized:
        return "destinationSchool"
    if "拟录取学院" in normalized:
        return "destinationCollege"
    if "拟录取专业" in normalized:
        return "destinationMajor"
    if normalized == "专业代码":
        return "majorCode"
    return ""


def find_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = [clean(value) for value in row]
        if not any("拟录取" in header for header in headers):
            continue
        mapping: dict[str, int] = {}
        for col_index, header in enumerate(headers):
            field = field_for_header(header)
            if field:
                mapping[field] = col_index
        if {"undergraduateMajor", "destinationSchool"}.issubset(mapping):
            return row_index, mapping
    raise ValueError(f"未找到表头：{ws.title}")


def cell(row: tuple, mapping: dict[str, int], field: str) -> object:
    index = mapping.get(field)
    if index is None or index >= len(row):
        return ""
    return row[index]


def import_file(path: Path) -> list[dict]:
    cohort, class_year, admission_year = cohort_from_file(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []

    for ws in wb.worksheets:
        header_row, mapping = find_header_row(ws)
        for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            undergraduate_major = clean(cell(row, mapping, "undergraduateMajor"))
            destination_school = strip_code_prefix(clean(cell(row, mapping, "destinationSchool")))
            if not undergraduate_major or not destination_school or "保研交流群" in undergraduate_major:
                continue

            undergraduate_college = clean(cell(row, mapping, "undergraduateCollege"))
            destination_college = strip_code_prefix(clean(cell(row, mapping, "destinationCollege")))
            destination_major = strip_code_prefix(clean(cell(row, mapping, "destinationMajor")))
            recommendation_type = clean(cell(row, mapping, "recommendationType"))
            major_code = format_code(cell(row, mapping, "majorCode"))

            records.append(
                {
                    "id": stable_id(
                        "admission",
                        path.name,
                        ws.title,
                        str(row_index),
                        undergraduate_major,
                        destination_school,
                        destination_college,
                        destination_major,
                        major_code,
                    ),
                    "cohort": cohort,
                    "classYear": class_year,
                    "admissionYear": admission_year,
                    "undergraduateCollege": undergraduate_college,
                    "undergraduateMajor": undergraduate_major,
                    "recommendationType": recommendation_type,
                    "destinationSchool": destination_school,
                    "destinationCollege": destination_college,
                    "destinationMajor": destination_major,
                    "majorCode": major_code,
                    "source": GENERATED_SOURCE,
                    "sourceFile": path.name,
                    "sourceSheet": ws.title,
                    "sourceRow": row_index,
                }
            )

    return records


def load_data(path: Path) -> dict:
    if not path.exists():
        return {"colleges": [], "mentors": [], "intentions": [], "admissions": []}
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    return {
        "colleges": data.get("colleges") if isinstance(data.get("colleges"), list) else [],
        "mentors": data.get("mentors") if isinstance(data.get("mentors"), list) else [],
        "intentions": data.get("intentions") if isinstance(data.get("intentions"), list) else [],
        "admissions": data.get("admissions") if isinstance(data.get("admissions"), list) else [],
    }


def write_data(data_path: Path, records: list[dict]) -> dict:
    data = load_data(data_path)
    data["admissions"] = records
    data_path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {
        "totalColleges": len(data["colleges"]),
        "totalMentors": len(data["mentors"]),
        "totalIntentions": len(data["intentions"]),
        "totalAdmissions": len(data["admissions"]),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import historical admission destinations into website data.json.")
    parser.add_argument("--source-dir", type=Path, default=ROOT / "src")
    parser.add_argument("--data", type=Path, default=ROOT / "data.json")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    records: list[dict] = []
    seen_hashes: set[str] = set()
    skipped_duplicates: list[str] = []
    files = sorted(args.source_dir.glob(SOURCE_GLOB))

    for path in files:
        digest = file_digest(path)
        if digest in seen_hashes:
            skipped_duplicates.append(path.name)
            continue
        seen_hashes.add(digest)
        records.extend(import_file(path))

    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    for record in records:
        record["createdAt"] = now
        record["updatedAt"] = now

    result = {
        "sourceFiles": len(files),
        "usedFiles": len(files) - len(skipped_duplicates),
        "skippedDuplicateFiles": skipped_duplicates,
        "generatedAdmissions": len(records),
        "topDestinationSchools": Counter(record["destinationSchool"] for record in records).most_common(20),
        "topUndergraduateMajors": Counter(record["undergraduateMajor"] for record in records).most_common(20),
    }
    if not args.dry_run:
        result.update(write_data(args.data, records))

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
